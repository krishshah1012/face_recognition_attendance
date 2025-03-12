from django.shortcuts import render,redirect
from .models import Profile,LastFace
from .forms import ProfileForm
from django.db.models import Q
from django.http import HttpResponse
import cv2
import os
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import winsound
from django.utils import timezone
from mtcnn import MTCNN 
import face_recognition
import numpy as np
import pickle
import openpyxl
from datetime import datetime
# Create your views here.

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def index(request):
    scanned = LastFace.objects.all().order_by('date').reverse()
    present = Profile.objects.filter(present=True).order_by('updated').reverse()
    absent = Profile.objects.filter(present=False).order_by('shift')
    
    all_profiles = list(present) + list(absent)
    
    context = {
        'scanned': scanned,
        'present': present,
        'all_profiles':all_profiles
    }
    return render(request, 'core/index.html', context)

def profiles(request):
    profiles = Profile.objects.all()
    context = {
        'profiles': profiles
    }
    return render(request, 'core/profiles.html', context)

from django.shortcuts import render
from .models import LastFace, Profile

# def details(request):
#     try:
#         last_face = LastFace.objects.last()
#         recognized_name = last_face.last_face

#         # Search for the profile with the recognized name
#         profile = Profile.objects.get(Q(first_name__icontains=recognized_name.split()[0]) &
#                                       Q(last_name__icontains=recognized_name.split()[1]))
#     except Profile.DoesNotExist:
#         last_face = None
#         profile = None

#     context = {
#         'profile': profile,
#         'last_face': last_face
#     }
#     return render(request, 'core/details.html', context)

# def details(request):
#     last_face = LastFace.objects.last()
#     profile = None  # Initialize profile as None

#     if last_face is not None:
#         recognized_name = last_face.last_face

#         try:
#             # Search for the profile with the recognized name
#             profile = Profile.objects.get(Q(first_name__icontains=recognized_name.split()[0]) &
#                                           Q(last_name__icontains=recognized_name.split()[1]))
#         except Profile.DoesNotExist:
#             pass  # Profile not found, profile remains None

#     context = {
#         'profile': profile,
#         'last_face': last_face
#     }
#     return render(request, 'core/details.html', context)


def add_profile(request):
    form = ProfileForm()
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('profiles')
    context = {'form':form}
    return render(request,'core/add_profile.html',context)

def edit_profile(request,id):
    profile = Profile.objects.get(id=id)
    form = ProfileForm(instance = profile)
    if request.method == 'POST':
        form = ProfileForm(request.POST,request.FILES,instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profiles')
    context = {'form':form}
    return render(request,'core/add_profile.html',context)

def delete_profile(request,id):
    profile = Profile.objects.get(id=id)
    profile.delete()
    return redirect('profiles')

def reset(request):
    current_day, current_time = get_current_day_and_time()
    initialize_attendance(current_day)
    profiles = Profile.objects.all()
    for profile in profiles:
        if profile.present == True:
            profile.present = False
            profile.save()
        else:
            pass
    return redirect('index')

def clear_history(request):
    history = LastFace.objects.all()
    history.delete()
    return redirect('index')


def get_current_day_and_time():
    now = datetime.now()
    current_day = now.strftime("%A")  # Get the full name of the day (e.g., Monday)
    current_time = now.strftime("%H:%M")  # Get the current time in HH:MM format
    return current_day, current_time

def initialize_attendance(day):
    # Load the workbook
    workbook = openpyxl.load_workbook('attendance.xlsx')
    # Get the current day's sheet
    sheet = workbook[day]
    # Iterate through all cells (except the first column which contains names)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            cell.value = "A"
    # Save the workbook
    workbook.save('attendance.xlsx')

def mark_attendance(day, time, name):
    # Load the workbook
    workbook = openpyxl.load_workbook('attendance.xlsx')

    # Get the current day's sheet
    sheet = workbook[day]

    # Find the row corresponding to the name
    for row in range(2, sheet.max_row + 1):  # Assuming names start from the 2nd row
        if sheet.cell(row=row, column=1).value == name:  # Assuming names are in the first column
            # Get the column corresponding to the current time
            if "08:00" <= time < "09:00":
                column = 2
            elif "09:00" <= time < "10:00":
                column = 3
            elif "10:00" <= time < "11:00":
                column = 4
            elif "11:00" <= time < "12:00":
                column = 5
            elif "12:00" <= time < "13:00":
                column = 6
            elif "13:00" <= time < "14:00":
                column = 7
            else:  # Attendance after 2:00 PM
                column = sheet.max_column

            # Mark attendance in the determined column
            sheet.cell(row=row, column=column).value = "P"
            break

    # Save the workbook
    workbook.save('attendance.xlsx')





# Function to load or train the face recognition model
def load_or_train_face_model(data_folder, model_file="face_model.pkl"):
    model_path = os.path.join(data_folder, model_file)

    if os.path.exists(model_path):
        # Load the model from the file
        with open(model_path, 'rb') as file:
            known_face_encodings, known_face_names = pickle.load(file)
    else:
        # Train the model and save it to a new file
        known_face_encodings, known_face_names = train_face_model(data_folder)
        with open(model_path, 'wb') as file:
            pickle.dump((known_face_encodings, known_face_names), file)

    return known_face_encodings, known_face_names

# Function to load images and train the face recognition model
def train_face_model(data_folder):
    known_face_encodings = []
    known_face_names = []

    for person_folder in os.listdir(data_folder):
        person_path = os.path.join(data_folder, person_folder)

        if os.path.isdir(person_path):
            for filename in os.listdir(person_path):
                image_path = os.path.join(person_path, filename)

                image = face_recognition.load_image_file(image_path)
                # Detect faces in the image
                face_locations = face_recognition.face_locations(image)

                if face_locations:
                    # If at least one face is detected, use the first one
                    face_encoding = face_recognition.face_encodings(image, [face_locations[0]])[0]
                    known_face_encodings.append(face_encoding)
                    known_face_names.append(person_folder)
                else:
                    print(f"No faces found in {image_path}")

    return known_face_encodings, known_face_names

# Function to detect and recognize faces in real-time from the webcam
def scan(request):
    current_day, current_time = get_current_day_and_time()
    
    sound = os.path.join(BASE_DIR, 'core', 'sound', 'beep.wav')
    video_capture = cv2.VideoCapture(0)
    recognized_employees = []

    data_folder = "faces"
    model_file = "face_model.pkl"
    profiles = Profile.objects.all()
    for profile in profiles:
        profile.present = False

    known_face_encodings, known_face_names = load_or_train_face_model(data_folder, model_file)
    last_face = ""

    while True:
        # Capture video frames
        ret, frame = video_capture.read()

        # Find all face locations and face encodings in the current frame
        face_locations = face_recognition.face_locations(frame)
        face_encodings = face_recognition.face_encodings(frame, face_locations)

        # Loop through each face found in the frame
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # Check if the face matches any known faces
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # If a match is found, use the name associated with the matched face
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                try:

                    profile = Profile.objects.get(first_name=name)
                    face_name = f'{profile.first_name} {profile.last_name}'
                    recognized_employees.append({
                        'name': face_name,
                        'profession': profile.profession,
                        'ranking': profile.ranking
                    })
                    mark_attendance(current_day, current_time, name)
                    profile.present = True
                    profile.updated = timezone.now()
                    profile.save()
                    
                    
                    
                    if last_face != face_name:
                        last_face = face_name
                        last_face_obj = LastFace(last_face=last_face)
                        last_face_obj.save()
                except Profile.DoesNotExist:
                    print(f"Profile with first name '{name}' does not exist in the database")


            # Draw a rectangle around the face and display the name
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

        # Display the resulting frame
        cv2.imshow('Video', frame)

        # Break the loop if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord('q')  :
            break

    # Release the video capture object
    video_capture.release()

    # Close all OpenCV windows
    cv2.destroyAllWindows()
    
def ajax(request):
    last_face_obj = LastFace.objects.last()
    last_face_value = last_face_obj.last_face if last_face_obj else None
    context = {
        'last_face': last_face_value
    }
    return JsonResponse (context)